from pathlib import Path
import pandas as pd
import re
import json
from rapidfuzz import process, fuzz
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ==========================
# الإعدادات
# ==========================
def merge_brands():
    MIN_BRANDS = 2

    folder = Path("output")

    with open("aliases.json", "r", encoding="utf-8") as f:
        ALIASES = json.load(f)

    REMOVE_WORDS = [
        "inspired by",
        "by",
        "louis vuitton",
        "lv",
        "versace",
        "xerjoff",
        "parfums de marly",
        "maison francis kurkdjian",
        "mfk",
        "jean paul gaultier",
        "guerlain",
        "kilian",
        "dior",
        "ysl",
        "bvlgari",
        "creed",
        "maison crivelli",
        "profumo di firenze",
        "argos",
    ]


    def normalize(name):
        if pd.isna(name):
            return ""

        name = str(name).lower()

        name = (
            name.replace("’", "'")
                .replace("‘", "'")
                .replace("–", " ")
                .replace("-", " ")
        )

        for word in REMOVE_WORDS:
            name = name.replace(word, "")

        name = re.sub(r"[^\w\s]", " ", name)
        name = " ".join(name.split())

        # لو موجود في aliases استخدم الاسم الموحد
        return ALIASES.get(name, name.title())


    # ==========================
    # قراءة الملفات
    # ==========================

    all_rows = []

    for file in folder.glob("*.xlsx"):

        if file.stem == "Price Comparison":
            continue

        brand = file.stem

        df = pd.read_excel(file)

        df.columns = df.columns.str.strip()

        df = df[["Original Perfume", "السعر"]].copy()

        df["Brand"] = brand

        df["Key"] = df["Original Perfume"].apply(normalize)

        all_rows.append(df)

    master = pd.concat(all_rows, ignore_index=True)

    # ==========================
    # دمج الأسماء المتشابهة
    # ==========================

    unique = []
    mapping = {}

    for key in master["Key"].unique():

        if not unique:
            unique.append(key)
            mapping[key] = key
            continue

        match = process.extractOne(
            key,
            unique,
            scorer=fuzz.token_sort_ratio
        )

        if match and match[1] >= 92:
            mapping[key] = match[0]
        else:
            unique.append(key)
            mapping[key] = key

    master["Key"] = master["Key"].map(mapping)

    # ==========================
    # Pivot
    # ==========================

    comparison = master.pivot_table(
        index="Key",
        columns="Brand",
        values="السعر",
        aggfunc="first"
    )

    comparison["عدد البراندات"] = comparison.notna().sum(axis=1)

    comparison = comparison[
        comparison["عدد البراندات"] >= MIN_BRANDS
    ]

    brand_cols = comparison.columns.drop("عدد البراندات")

    comparison["أقل سعر"] = comparison[brand_cols].min(axis=1)
    comparison["أعلى سعر"] = comparison[brand_cols].max(axis=1)
    comparison["فرق السعر"] = comparison["أعلى سعر"] - comparison["أقل سعر"]

    comparison = comparison.reset_index()

    comparison = comparison.rename(
        columns={
            "Key": "Original Perfume"
        }
    )

    comparison = comparison.sort_values(
        by=["عدد البراندات", "أقل سعر", "Original Perfume"],
        ascending=[False, True, True]
    )

    # ترتيب الأعمدة
    fixed = [
        "Original Perfume",
        "عدد البراندات",
        "أقل سعر",
        "أعلى سعر",
        "فرق السعر"
    ]

    brands = sorted(
        [
            c for c in comparison.columns
            if c not in fixed
        ]
    )

    comparison = comparison[
        fixed + brands
    ]

    excel_file = folder / "Price Comparison.xlsx"

    comparison.to_excel(
        excel_file,
        index=False
    )

    # ==========================================
    # تنسيق ملف الإكسيل
    # ==========================================

    wb = load_workbook(excel_file)
    ws = wb.active

    # تنسيق الهيدر
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Freeze Header
    ws.freeze_panes = "A2"

    # Filter
    ws.auto_filter.ref = ws.dimensions

    # Auto Width
    for column in ws.columns:
        max_length = 0
        letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[letter].width = max_length + 3

    # ==========================================
    # تلوين أقل وأعلى سعر
    # ==========================================

    green = PatternFill(fill_type="solid", fgColor="C6EFCE")
    red = PatternFill(fill_type="solid", fgColor="FFC7CE")

    headers = [c.value for c in ws[1]]

    brand_columns = []

    for i, h in enumerate(headers, start=1):
        if h in ["COZY", "IL_VANTO", "Lail", "MIXI"]:
            brand_columns.append(i)

    for row in range(2, ws.max_row + 1):

        prices = []

        for col in brand_columns:

            value = ws.cell(row, col).value

            if isinstance(value, (int, float)):
                prices.append((col, value))

        if len(prices) < 2:
            continue

        min_price = min(v for _, v in prices)
        max_price = max(v for _, v in prices)

        for col, value in prices:

            if value == min_price:
                ws.cell(row, col).fill = green

            if value == max_price:
                ws.cell(row, col).fill = red

    wb.save(excel_file)

    print("Done ✔")
    
if __name__ == "__main__":
    merge_brands()